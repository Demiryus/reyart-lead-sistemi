"""
Reyart Lead Enricher — Adım 2 (hızlı, paralel)
Her firma için:
  - Bing HTML arama → website bul
  - Website /contact sayfasından telefon/e-posta çek
  - LinkedIn URL bul
Çıktı: output/leads.xlsx
"""

import json
import re
import time
import random
import logging
import subprocess
import sys
import urllib3
from difflib import SequenceMatcher
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# Mutlak yol: webapp gibi farklı çalışma dizininden import edilse de doğru yeri bulur
OUTPUT_DIR = Path(__file__).parent / "output"
COMPANIES_FILE = OUTPUT_DIR / "companies.json"
EXCEL_FILE = OUTPUT_DIR / "leads.xlsx"

# Enricher'ın sahiplendiği alanlar — kaydetme SADECE bunları birleştirir.
# Eskiden tüm listeyi olduğu gibi diske yazıyordu; saatlerce süren bir koşu
# sırasında webapp'ten yapılan durum/not düzenlemelerini ve temizlik
# script'lerinin sildiği kayıtları geri getirip eziyordu (2026-07 bug'ı).
ENRICH_FIELDS = ("website", "phone", "email", "linkedin",
                 "site_dogrulama", "tel_dogrulama")


def save_progress(enriched: list[dict], backup: bool = False) -> list[dict]:
    """Diskteki GÜNCEL companies.json'ı yeniden okuyup, zenginleştirilen
    kayıtların iletişim alanlarını id bazında (sadece diskte boşsa) doldurur
    ve atomik kaydeder. Diskten silinmiş kayıtları geri EKLEMEZ; webapp'te
    o sırada değişen status/note/takip alanlarına dokunmaz."""
    import datastore as ds
    with ds._WRITE_LOCK:
        disk = ds.load()
        by_id = {c.get("id"): c for c in disk if c.get("id")}
        merged = 0
        for e in enriched:
            eid = e.get("id") or ds.make_id(e)
            t = by_id.get(eid)
            if t is None:
                continue  # diskten silinmiş (ör. junk temizliği) — geri getirme
            for f in ENRICH_FIELDS:
                v = e.get(f, "")
                if v and not t.get(f):
                    t[f] = v
                    merged += 1
        ds.save(disk, backup=backup)
    return disk

# Google Maps fallback — telefon eksik kalan firmalar için.
# Worker artık projenin içinde (maps_worker.py); eski masaüstü kopyası
# sadece geriye dönük yedek olarak denenir.
MAPS_WORKER = Path(__file__).parent / "maps_worker.py"
if not MAPS_WORKER.exists():
    MAPS_WORKER = Path(r"C:\Users\Home\Desktop\11.06.26\_cf_worker.py")
MAPS_HEADLESS = True
MAPS_DELAY = 1.5

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "tr-TR,tr;q=0.9,en-US;q=0.8,en;q=0.7",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

PHONE_RE = re.compile(
    r"(?:\+90[\s\-.]?|0)"
    r"(?:(?:2\d{2}|[35]\d{2}|444|850)[\s\-.]?\d{3}[\s\-.]?\d{2}[\s\-.]?\d{2})",
)
# Yabancı firmaların sitelerindeki uluslararası numaralar (+49, +39, +86...).
# +90 hariç (onu TR regex'i daha sıkı yakalar). En az 8, en çok 15 rakam (E.164).
INTL_PHONE_RE = re.compile(
    r"\+(?!90)(\d{1,3})[\s\-.]?\(?\d{1,4}\)?(?:[\s\-.]?\d{1,4}){2,5}"
)
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")

CONTACT_PATHS = [
    "/contact", "/contact-us", "/iletisim",
    "/tr/iletisim", "/en/contact", "/about/contact",
    "/hakkimizda/iletisim", "/tr/contact-us",
]

TIMEOUT = 10
MAX_WORKERS = 6

# ── İsim eşleştirme: bulunan site/LinkedIn/telefon gerçekten bu firmaya mı ait? ──
# "resmi site" aramasında ilk sonucu körlemesine almak yanlış firmanın telefonunu
# getiriyordu (özellikle kısa/özel isimli firmalarda). Kabul etmeden önce sonucun
# içeriğinde firma adının anlamlı kelimeleri geçiyor mu diye kontrol ediyoruz.
LEGAL_STOPWORDS = {
    "ve", "a.s.", "as", "a.ş.", "ltd", "ltd.", "sti", "sti.", "şti", "şti.",
    "san", "san.", "tic", "tic.", "dis", "dış", "ith", "ihr", "ithalat",
    "ihracat", "pazarlama", "paz", "paz.", "co", "co.", "inc", "inc.",
    "gmbh", "kg", "srl", "s.r.l", "spa", "s.p.a", "ag", "sa", "sirketi",
    "şirketi", "anonim", "limited", "sanayi", "ticaret", "endustri",
    "endüstri", "turkiye", "türkiye", "turkey", "and", "the",
}

# Token-overlap eşiği tutmayan ama yazım farkı/şube-marka varyasyonu
# ("ASD Laminat" vs "ASD Laminat Canada") olan adaylar için ek fuzzy sinyal.
FUZZY_RATIO_THRESHOLD = 0.6


def _fold_tr(s: str) -> str:
    table = str.maketrans({
        "İ": "i", "I": "i", "ı": "i", "Ş": "s", "ş": "s", "Ğ": "g", "ğ": "g",
        "Ü": "u", "ü": "u", "Ö": "o", "ö": "o", "Ç": "c", "ç": "c",
    })
    return s.translate(table)


def sig_words(name: str) -> set[str]:
    """Firma adından ayırt edici kelimeleri çıkarır (hukuki ekleri atar)."""
    # ÖNEMLİ: fold_tr önce, .lower() sonra — Python'da "İ".lower() bir
    # combining-dot karakteri (U+0307) ekliyor ve "tic" gibi kelimeleri
    # stopword listesiyle eşleşmez hale getiriyordu.
    s = _fold_tr(name).lower()
    s = re.sub(r"[^\w\s]", " ", s)
    return {w for w in s.split() if w not in LEGAL_STOPWORDS and len(w) > 1}


def _alnum_norm(s: str) -> str:
    """Fuzzy karşılaştırma için boşluk/noktalama atılmış, TR karakterleri sadeleştirilmiş hali."""
    return re.sub(r"[^a-z0-9]", "", _fold_tr(s).lower())


def content_matches(html: str, url: str, name: str) -> bool:
    """Bulunan sayfa/domain gerçekten bu firmaya mı ait, kelime örtüşmesiyle kontrol et.
    Token-overlap eşiği tutmazsa domain/isim arasındaki fuzzy benzerlik de ek sinyal
    olarak denenir (typo/şube-marka varyasyonlarını yakalamak için)."""
    wanted = sig_words(name)
    domain = re.sub(r"^https?://(www\.)?", "", url).split("/")[0]
    if not wanted:
        return True
    haystack = _fold_tr(domain + " " + html[:3000]).lower()
    hits = sum(1 for w in wanted if w in haystack)
    threshold = max(1, len(wanted) // 2)
    token_ok = hits >= threshold
    ratio = SequenceMatcher(None, _alnum_norm(name), _alnum_norm(domain)).ratio()
    fuzzy_ok = ratio >= FUZZY_RATIO_THRESHOLD
    if not (token_ok or fuzzy_ok):
        log.debug(
            "REDDEDILDI '%s' için aday %s: kelime örtüşmesi %d/%d (eşik %d), fuzzy benzerlik %.2f",
            name, url, hits, len(wanted), threshold, ratio,
        )
        return False
    return True


def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update(HEADERS)
    s.verify = False
    return s


def _get(session: requests.Session, url: str) -> str:
    try:
        r = session.get(url, timeout=TIMEOUT, allow_redirects=True)
        if r.status_code == 200:
            return r.text
    except Exception:
        pass
    return ""


def extract_contacts(html: str) -> tuple[str, str, str]:
    phones = PHONE_RE.findall(html)
    emails = EMAIL_RE.findall(html)
    li_matches = re.findall(
        r'https?://(?:www\.)?linkedin\.com/company/[^\s"\'<>&/][^\s"\'<>&]*',
        html,
    )
    phone = ""
    if phones:
        raw = re.sub(r"[\s\-.]", "", phones[0])
        phone = raw if raw.startswith("+") else ("+9" + raw if raw.startswith("0") else "+90" + raw)
    else:
        # TR numarası yoksa uluslararası dene (yabancı firma siteleri)
        m = INTL_PHONE_RE.search(html)
        if m:
            raw = re.sub(r"[\s\-.()]", "", m.group(0))
            if 9 <= len(raw.lstrip("+")) <= 15:  # E.164 sınırı
                phone = raw

    junk = {"example", "test", "sample", "noreply", "no-reply", "info@example"}
    email = next((e for e in emails if not any(j in e.lower() for j in junk)), "")
    linkedin = li_matches[0].rstrip("/") if li_matches else ""
    return phone, email, linkedin


def bing_search_candidates(session: requests.Session, query: str, limit: int = 5) -> list[str]:
    """Bing HTML arama, ilk N geçerli sonucu döndürür (tek sonuca güvenmek yanlış eşleşme getiriyordu)."""
    url = f"https://www.bing.com/search?q={requests.utils.quote(query)}&setlang=tr"
    html = _get(session, url)
    if not html:
        return []
    soup = BeautifulSoup(html, "lxml")

    links = []
    for a in soup.select(".b_algo h2 a, .b_title a"):
        href = a.get("href", "")
        if href.startswith("http") and "bing.com" not in href and "microsoft.com" not in href:
            links.append(href)
            if len(links) >= limit:
                break
    return links


def bing_search(session: requests.Session, query: str) -> str:
    """Geriye dönük uyumluluk için: ilk sonucu döndürür."""
    candidates = bing_search_candidates(session, query, limit=1)
    return candidates[0] if candidates else ""


def guess_website(name: str) -> list[str]:
    """Firma adından olası URL'ler üret."""
    clean = re.sub(r"[^\w\s]", "", name.lower())
    words = clean.split()

    candidates = []
    if words:
        slug = words[0]
        # İlk kelime + com/com.tr
        candidates += [
            f"https://www.{slug}.com.tr",
            f"https://www.{slug}.com",
        ]
        if len(words) >= 2:
            slug2 = words[0] + words[1]
            candidates += [
                f"https://www.{slug2}.com.tr",
                f"https://www.{slug2}.com",
            ]
    return candidates


def find_website(session: requests.Session, name: str) -> str:
    # 1. Bing arama — birden fazla adayı dene, ilk sonucu körlemesine kabul etme
    time.sleep(random.uniform(0.5, 1.2))
    for url in bing_search_candidates(session, f"{name} resmi site OR official website"):
        html = _get(session, url)
        if html and len(html) > 500 and content_matches(html, url, name):
            return url
        time.sleep(0.2)

    # 2. Tahmin et ve dene — burada da içerik firma adıyla örtüşmeli
    for url in guess_website(name):
        html = _get(session, url)
        if html and len(html) > 500 and content_matches(html, url, name):
            return url
        time.sleep(0.2)

    return ""


def scrape_website(session: requests.Session, website: str) -> tuple[str, str, str]:
    phone, email, linkedin = "", "", ""
    base = website.rstrip("/")
    if not base.startswith("http"):
        base = "https://" + base

    pages = [base] + [base + p for p in CONTACT_PATHS]
    for page_url in pages:
        html = _get(session, page_url)
        if not html:
            continue
        p, e, li = extract_contacts(html)
        phone = phone or p
        email = email or e
        linkedin = linkedin or li
        if phone and email and linkedin:
            break
        time.sleep(0.2)

    return phone, email, linkedin


def find_linkedin(session: requests.Session, name: str) -> str:
    time.sleep(random.uniform(0.4, 0.9))
    candidates = [
        c for c in bing_search_candidates(session, f"{name} linkedin.com/company")
        if "linkedin.com/company" in c
    ]
    # Bing HTML içinden direkt link ara (ilk aramada bulunamadıysa)
    url = f"https://www.bing.com/search?q={requests.utils.quote(name + ' site:linkedin.com/company')}"
    html = _get(session, url)
    if html:
        candidates += re.findall(
            r'https?://(?:www\.)?linkedin\.com/company/[^\s"\'<>&?]+',
            html,
        )

    wanted = sig_words(name)
    for li in candidates:
        li = li.rstrip("/")
        slug = li.split("/company/")[-1].replace("-", " ")
        haystack = _fold_tr(slug).lower()
        hits = sum(1 for w in wanted if w in haystack)
        threshold = max(1, len(wanted) // 2)
        token_ok = (not wanted) or (hits >= threshold)
        ratio = SequenceMatcher(None, _alnum_norm(name), _alnum_norm(slug)).ratio()
        fuzzy_ok = ratio >= FUZZY_RATIO_THRESHOLD
        if token_ok or fuzzy_ok:
            return li
        log.debug(
            "REDDEDILDI '%s' için LinkedIn adayı %s: kelime örtüşmesi %d/%d (eşik %d), fuzzy benzerlik %.2f",
            name, li, hits, len(wanted), threshold, ratio,
        )
    return ""


def enrich_one(company: dict) -> dict:
    session = make_session()
    name = company["name"]

    # Website bul (find_website artık sadece isimle eşleşen siteyi döndürür)
    if not company.get("website"):
        site = find_website(session, name)
        company["website"] = site
        company["site_dogrulama"] = "✅ Site adla eşleşti" if site else "❌ Eşleşen site bulunamadı"

    # Website'den iletişim bilgisi çek
    phone, email, linkedin = "", "", ""
    if company["website"]:
        phone, email, linkedin = scrape_website(session, company["website"])

    if phone:
        company["phone"] = phone
    if email:
        company["email"] = email
    if linkedin:
        company["linkedin"] = linkedin

    # LinkedIn hâlâ boşsa ara
    if not company["linkedin"]:
        company["linkedin"] = find_linkedin(session, name)

    return company


# ── Google Maps fallback ──────────────────────────────────────────────────────

def run_maps_fallback(companies: list[dict]) -> list[dict]:
    """Telefonu eksik olan firmaları MapsScraperProv1 worker'ına gönder."""
    if not MAPS_WORKER.exists():
        log.warning("Maps worker bulunamadı (%s), atlanıyor", MAPS_WORKER)
        return companies

    eksikler = [c for c in companies if not c.get("phone")]
    if not eksikler:
        log.info("Tüm firmaların telefonu var, Maps fallback gerekmiyor")
        return companies

    log.info("=" * 55)
    log.info("GOOGLE MAPS FALLBACK: %d firma için telefon aranıyor", len(eksikler))

    cfg_path = OUTPUT_DIR / "_maps_cfg.json"
    out_path = OUTPUT_DIR / "_maps_out.jsonl"
    if out_path.exists():
        out_path.unlink()

    cfg_path.write_text(
        json.dumps({
            # Ülke bilgisiyle: worker yabancı firmalar için "Türkiye" eklemeden
            # arama yapabilsin (bkz. _cf_worker.py main()).
            "companies": [
                {"name": c["name"], "country": c.get("country", "")} for c in eksikler
            ],
            # Geriye dönük uyumluluk: eski worker sürümleri "names" bekliyor.
            "names": [c["name"] for c in eksikler],
            "headless": MAPS_HEADLESS,
            "delay": MAPS_DELAY,
        }, ensure_ascii=False),
        encoding="utf-8",
    )

    try:
        proc = subprocess.run(
            [sys.executable, str(MAPS_WORKER), str(cfg_path), str(out_path)],
            capture_output=True,
            text=True,
            timeout=3600,
        )
        if proc.returncode != 0:
            log.warning("Maps worker hata kodu %d: %s", proc.returncode, proc.stderr[:300])
    except subprocess.TimeoutExpired:
        log.warning("Maps worker timeout oldu, kısmi sonuç okunuyor")
    except Exception as e:
        log.error("Maps worker çalıştırılamadı: %s", e)
        return companies

    # JSONL sonuçlarını oku
    if not out_path.exists():
        log.warning("Maps çıktısı yok")
        return companies

    by_name: dict[str, dict] = {}
    with open(out_path, encoding="utf-8") as f:
        for line in f:
            try:
                obj = json.loads(line)
                if obj.get("type") == "result":
                    by_name[obj["isim"]] = obj
            except Exception:
                continue

    # Firma listesini güncelle
    found = 0
    for c in companies:
        if c.get("phone"):
            continue
        result = by_name.get(c["name"])
        if not result:
            continue
        phones = result.get("telefonlar", [])
        if phones:
            raw = phones[0]
            c["phone"] = raw if raw.startswith("+") else "+9" + raw if raw.startswith("0") else "+90" + raw
            found += 1
            yer = result.get("eslesen_yer", "")
            c["tel_dogrulama"] = f"✅ Maps: {yer} eşleşti" if yer else "✅ Maps eşleşti"
            # Adres bilgisini Not kolonuna ekle
            kaynaklar = result.get("kaynaklar", [])
            if kaynaklar and not c.get("note"):
                c["note"] = "Maps: " + kaynaklar[0][:80]
        else:
            c["tel_dogrulama"] = "❌ Maps'te eşleşen yer bulunamadı"

    log.info("MAPS: %d eksik firmadan %d'sine telefon eklendi", len(eksikler), found)

    # Geçici dosyaları temizle
    try:
        cfg_path.unlink()
        out_path.unlink()
    except Exception:
        pass

    return companies


# ── Excel ─────────────────────────────────────────────────────────────────────


def compute_guven_seviyesi(c: dict) -> str:
    """Site (content_matches) VE telefon (Maps place_matches) bağımsız kaynaklardan
    doğrulanmışsa dual-source olarak 'Yüksek' işaretle. Alan adı: guven_seviyesi."""
    site_ok = str(c.get("site_dogrulama", "")).startswith("✅")
    tel_ok = str(c.get("tel_dogrulama", "")).startswith("✅")
    if site_ok and tel_ok:
        return "🟢 Yüksek"
    if site_ok or tel_ok:
        return "🟡 Orta"
    return "⚪ Belirsiz"
FILL_GOLD   = PatternFill("solid", fgColor="FFD700")
FILL_BLUE   = PatternFill("solid", fgColor="D9E8F6")
FILL_WHITE  = PatternFill("solid", fgColor="FFFFFF")
FILL_HEADER = PatternFill("solid", fgColor="2F4F8F")
FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_NORMAL = Font(name="Calibri", size=10)
THIN   = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Doğrulama/Güven Seviyesi hücreleri için durum renkleri — satırın öncelik
# rengini (FILL_GOLD/BLUE/WHITE) EZMEZ, sadece bu iki kolonun hücresine
# override olarak uygulanır (bkz. save_excel).
FILL_CONFIRM_GREEN  = PatternFill("solid", fgColor="C6EFCE")  # 🟢 Yüksek / tam ✅
FILL_CONFIRM_YELLOW = PatternFill("solid", fgColor="FFEB9C")  # 🟡 Orta / karışık ✅+❌
FILL_CONFIRM_RED    = PatternFill("solid", fgColor="FFC7CE")  # ⚪ Belirsiz / ❌ / henüz kontrol yok


def _dogrulama_fill(text: str) -> PatternFill:
    """Doğrulama hücresi metnine göre durum rengi (✅/❌ karışık gelebilir, " | " ile birleşik)."""
    has_check = "✅" in text
    has_cross = "❌" in text
    if has_check and has_cross:
        return FILL_CONFIRM_YELLOW   # biri doğrulandı biri doğrulanamadı → orta
    if has_check:
        return FILL_CONFIRM_GREEN
    return FILL_CONFIRM_RED          # ❌ veya "➖ Henüz kontrol edilmedi"


def _guven_fill(text: str) -> PatternFill:
    """Güven Seviyesi hücresi için 3 sabit değere göre durum rengi."""
    if "🟢" in text:
        return FILL_CONFIRM_GREEN
    if "🟡" in text:
        return FILL_CONFIRM_YELLOW
    return FILL_CONFIRM_RED          # ⚪ Belirsiz


# Her fuarda olmayabilecek alanlar — sadece o fuarın firmalarında gerçekten
# dolu olan varsa sheet'e kolon olarak eklenir (örn. Intermob'da Salon/Stant
# var, KOMATEK/WIN EURASIA'da yok — boş kolon eklemeye gerek yok).
OPTIONAL_COLUMNS = [
    ("Salon", 10, "hall"), ("Stant", 12, "stand"), ("Detay URL", 45, "detail_url"),
]


def _invalid_sheet_chars(name: str) -> str:
    for ch in r"[]:*?/\\":
        name = name.replace(ch, " ")
    return name[:31]


def _write_fair_sheet(wb, fair_name: str, companies: list[dict], title: str | None = None,
                      include_fair_col: bool = False):
    """Tek fuarın (veya birleşik grubun) firmalarını uygun kolonlarla sheet'e yazar."""
    extra_cols = [
        (label, width, key) for label, width, key in OPTIONAL_COLUMNS
        if any(c.get(key) for c in companies)
    ]
    base_cols = [
        ("Firma Adı", 35, "name"), ("Menşei", 15, "country"), ("Sektör", 20, "sector"),
        ("Website", 30, "website"), ("Telefon", 18, "phone"), ("E-posta", 28, "email"),
        ("LinkedIn", 35, "linkedin"),
    ]
    if include_fair_col:
        base_cols.insert(1, ("Fuar", 32, "fair"))
    tail_cols = [
        ("Öncelik", 10, "priority"), ("Sıcaklık", 13, "sicaklik"),
        ("Fuar Tarihi", 12, "fuar_tarihi"), ("Satış Skoru", 11, "satis_skoru"),
        ("Durum", 15, "status"), ("Takip", 11, "takip_tarihi"),
        ("Doğrulama", 40, None), ("Güven Seviyesi", 16, None), ("Not", 25, "note"),
    ]
    columns = base_cols + extra_cols + tail_cols
    dogrulama_col = next(i for i, c in enumerate(columns, start=1) if c[0] == "Doğrulama")
    guven_col = next(i for i, c in enumerate(columns, start=1) if c[0] == "Güven Seviyesi")

    ws = wb.create_sheet(title or _invalid_sheet_chars(fair_name))
    for col_idx, (col_name, col_width, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for row_idx, c in enumerate(companies, start=2):
        priority = c.get("priority", "⭐")
        fill = FILL_GOLD if priority == "⭐⭐⭐" else (FILL_BLUE if priority == "⭐⭐" else FILL_WHITE)
        dogrulama = " | ".join(
            filter(None, [c.get("site_dogrulama", ""), c.get("tel_dogrulama", "")])
        ) or "➖ Henüz kontrol edilmedi"

        for col_idx, (col_name, _, key) in enumerate(columns, start=1):
            if col_name == "Doğrulama":
                val = dogrulama
            elif col_name == "Güven Seviyesi":
                val = c["guven_seviyesi"]
            elif col_name == "Durum":
                val = c.get("status", "⬜ Aranmadı")
            else:
                val = c.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.font = FONT_NORMAL
            cell.alignment = Alignment(vertical="center")
            cell.border = BORDER

        # Doğrulama/Güven Seviyesi hücreleri: satırın öncelik rengini koru,
        # sadece bu iki hücreyi kendi durumuna göre override et.
        ws.cell(row=row_idx, column=dogrulama_col).fill = _dogrulama_fill(dogrulama)
        ws.cell(row=row_idx, column=guven_col).fill = _guven_fill(c["guven_seviyesi"])

    ws.auto_filter.ref = ws.dimensions


def save_excel(companies: list[dict], path=None):
    """path verilmezse EXCEL_FILE'a yazar. Dosya Excel'de açıkken (kilitliyken)
    export alınabilsin diye webapp geçici bir path verir."""
    target = Path(path) if path else EXCEL_FILE
    # Güven seviyesini önce hesapla — hem Özet hem fuar sheet'leri buna bakıyor.
    for c in companies:
        c["guven_seviyesi"] = compute_guven_seviyesi(c)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Özet sheet: fuar bazında sayım (ilk/açılış sekmesi) ─────────────────
    ws2 = wb.create_sheet("Özet")
    summary_cols = [
        ("Fuar", 26), ("Toplam Firma", 13), ("Website Var", 12),
        ("Telefon Var", 12), ("E-posta Var", 12), ("LinkedIn Var", 12),
        ("🟢 Yüksek Güven", 16), ("⚪ Belirsiz", 12),
    ]
    for col_idx, (col_name, col_width) in enumerate(summary_cols, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws2.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws2.row_dimensions[1].height = 22

    fairs: dict[str, dict] = {}
    for c in companies:
        fair_name = c.get("fair", "") or "(Fuar Belirtilmemiş)"
        s = fairs.setdefault(fair_name, {
            "total": 0, "web": 0, "tel": 0, "mail": 0, "li": 0,
            "yuksek": 0, "belirsiz": 0,
        })
        s["total"] += 1
        if c.get("website"):
            s["web"] += 1
        if c.get("phone"):
            s["tel"] += 1
        if c.get("email"):
            s["mail"] += 1
        if c.get("linkedin"):
            s["li"] += 1
        guven = c.get("guven_seviyesi", "")
        if "🟢" in guven:
            s["yuksek"] += 1
        if "⚪" in guven:
            s["belirsiz"] += 1

    row_idx = 2
    for fair_name in sorted(fairs.keys()):
        s = fairs[fair_name]
        row_vals = [
            fair_name, s["total"], s["web"], s["tel"], s["mail"], s["li"], s["yuksek"], s["belirsiz"],
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = Alignment(
                horizontal="left" if col_idx == 1 else "center", vertical="center"
            )
            cell.border = BORDER
        row_idx += 1

    total_row = [
        "TOPLAM",
        sum(s["total"] for s in fairs.values()),
        sum(s["web"] for s in fairs.values()),
        sum(s["tel"] for s in fairs.values()),
        sum(s["mail"] for s in fairs.values()),
        sum(s["li"] for s in fairs.values()),
        sum(s["yuksek"] for s in fairs.values()),
        sum(s["belirsiz"] for s in fairs.values()),
    ]
    for col_idx, val in enumerate(total_row, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.fill = FILL_GOLD
        cell.alignment = Alignment(
            horizontal="left" if col_idx == 1 else "center", vertical="center"
        )
        cell.border = BORDER
    ws2.freeze_panes = "A2"

    # ── Fuar sheet'leri: 150 fuar × 150 sheet kullanılamaz olur — sadece
    # MAJOR_SHEET_MIN ve üzeri firmalı fuarlar kendi sheet'ini alır, kalan
    # küçük fuarlar tek "Diğer Fuarlar" sheet'inde (Fuar kolonuyla) birleşir.
    MAJOR_SHEET_MIN = 50
    by_fair: dict[str, list[dict]] = {}
    for c in companies:
        by_fair.setdefault(c.get("fair", "") or "(Fuar Belirtilmemiş)", []).append(c)

    majors = sorted(
        [(f, cs) for f, cs in by_fair.items() if len(cs) >= MAJOR_SHEET_MIN],
        key=lambda kv: -len(kv[1]),
    )
    minors = [c for f, cs in by_fair.items() if len(cs) < MAJOR_SHEET_MIN for c in cs]

    seen_titles = {"Özet", "Diğer Fuarlar"}
    for fair_name, fair_companies in majors:
        title = _invalid_sheet_chars(fair_name)
        n = 2
        while title in seen_titles:  # 31 karaktere kırpılan uzun adlar çakışabilir
            title = f"{_invalid_sheet_chars(fair_name)[:28]}({n})"
            n += 1
        seen_titles.add(title)
        _write_fair_sheet(wb, fair_name, fair_companies, title=title)
    if minors:
        minors.sort(key=lambda c: (c.get("fuar_tarihi") or "9999", c.get("fair", "")))
        _write_fair_sheet(wb, "Diğer Fuarlar", minors, title="Diğer Fuarlar",
                          include_fair_col=True)

    wb.active = 0  # açılışta Özet sekmesi görünsün
    wb.save(target)
    log.info("Excel kaydedildi → %s (%d büyük fuar sheet'i + %s)",
             target, len(majors), "Diğer Fuarlar" if minors else "—")


# ── Türk firma sezgisi (--tr-only) ───────────────────────────────────────────
# Scrape edilen kayıtlarda country çoğu zaman boş → menşei tek başına yetmez;
# ünvan/isimden ayırt ederiz. Türk ünvan eki veya Türkçe karakter → Türk.
# Yabancı ünvan eki ya da Çin şehir/eyalet adı → yabancı. İkisi de yoksa
# (düz marka adı, ör. "Granio") Türk sayılır — TR fuarındaki markasız isimler
# ağırlıkla yerli.

_TR_MARKERS = ["ŞTİ", "ŞTI", "A.Ş", "SAN.", "TİC.", "TIC.", "SANAYİ", "SANAYI",
               "TİCARET", "TICARET", "LTD ŞTİ", "PAZARLAMA", "MÜMESSİLLİK"]
_FOREIGN_MARKERS = ["CO., LTD", "CO.,LTD", "CO. LTD", "CO.LTD", " INC", "GMBH",
                    "S.R.L", " SRL", " JSC", " PLC", " PVT", " LLC", " BV",
                    " SP. Z O.O", " SA ", " AG ", "TRADING CO", "IMPORT AND EXPORT",
                    "IMP&EXP", "IMP & EXP", " FACTORY", "MACHINERY CO", "TECHNOLOGY CO",
                    "INDUSTRIAL CO", "INDUSTRY CO", "ELECTRONICS CO", "HARDWARE CO",
                    "TEXTILE CO", "MATERIAL CO", "PRODUCTS CO", "EQUIPMENT CO",
                    "GUANGDONG", "FOSHAN", "GUANGZHOU", "SHENZHEN", "DONGGUAN",
                    "ZHEJIANG", "JIANGSU", "SHANDONG", "SHANGHAI", "NINGBO",
                    "YIWU", "HEBEI", "HENAN", "FUJIAN", "XIAMEN", "QINGDAO",
                    "HANGZHOU", "WENZHOU", "GANZHOU", "ANHUI", "SICHUAN", "TIANJIN",
                    "WUQIANG", "SHIJIAZHUANG", "ZHENGZHOU", "CHANGZHOU", "SUZHOU",
                    "TAIZHOU", "JINAN", "WUHAN", "CHONGQING", "XINXIANG", "BAODING",
                    "CANGZHOU", "JIAXING", "SHAOXING", "ZIBO", "WEIFANG", "LINYI"]
_TR_CHARS = set("çğışöüÇĞİŞÖÜ")

# Firma değil sayfa başlığı/menü artığı — zenginleştirme bunlara yanlış
# telefon/mail yapıştırmasın diye hedef dışı bırakılır.
_NOT_A_COMPANY = ["katilimci", "ziyaretci", "danisma kurulu", "sergilen",
                  "s.s.s", "sss", "hakkimizda", "hakkinda", "iletisim formu",
                  "duyuru", "basin bulteni", "stant basvuru", "e-bilet", "bilet al",
                  "yaklasan fuar", "fuar seciniz", "fuar takvimi", "fuar alani",
                  "online kayit", "ucretsiz davetiye", "davetiye al", "neden katil"]


def _looks_turkish(c: dict) -> bool:
    from scraper import _norm as _n
    country = (c.get("country") or "").strip().lower()
    if country and country not in ("türkiye", "turkiye", "turkey", "tr"):
        return False
    raw = c.get("name") or ""
    if any(p in _n(raw) for p in _NOT_A_COMPANY):
        return False
    name = raw.upper()
    if any(m in name for m in _TR_MARKERS) or any(ch in _TR_CHARS for ch in raw):
        return True
    if any(m in name for m in _FOREIGN_MARKERS):
        return False
    return True


# ── Ana akış ──────────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Reyart firma zenginleştirici")
    parser.add_argument(
        "--only-empty", action="store_true",
        help="Sadece web/telefon/email'i olmayan firmaları işle (mevcutları atla)",
    )
    parser.add_argument(
        "--no-maps", action="store_true",
        help="Google Maps fallback'ı atla (ayrı maps_only.py ile çalıştırılacaksa kullan)",
    )
    parser.add_argument(
        "--fair", default=None,
        help="Sadece belirtilen fuardaki firmaları işle (örn: 'Intermob 2026')",
    )
    parser.add_argument(
        "--window", nargs=2, metavar=("BASLANGIC", "BITIS"), default=None,
        help="Fuar tarihi bu aralıkta olan firmaları işle (fair_dates.json'dan; örn: --window 2026-10-01 2026-12-31)",
    )
    parser.add_argument(
        "--tr-only", action="store_true",
        help="Sadece Türk görünen firmaları işle (yabancı ünvan/Çin şehri adı taşıyanları atla)",
    )
    args = parser.parse_args()

    if not COMPANIES_FILE.exists():
        log.error("companies.json bulunamadı. Önce scraper.py çalıştırın.")
        return

    companies = json.loads(COMPANIES_FILE.read_text(encoding="utf-8"))
    total = len(companies)

    if args.fair:
        fair_idx = {i for i, c in enumerate(companies) if c.get("fair") == args.fair}
        if not fair_idx:
            log.error("'%s' fuarına ait firma bulunamadı.", args.fair)
            return
        log.info("Fuar filtresi: '%s' → %d firma", args.fair, len(fair_idx))
    else:
        fair_idx = set(range(total))

    if args.window:
        fd_file = Path(__file__).parent / "output" / "fair_dates.json"
        fd = json.loads(fd_file.read_text(encoding="utf-8")) if fd_file.exists() else {}
        lo, hi = args.window
        fair_idx &= {i for i, c in enumerate(companies)
                     if lo <= (fd.get(c.get("fair", "")) or "") <= hi}
        log.info("Tarih penceresi %s → %s: %d firma", lo, hi, len(fair_idx))

    if args.tr_only:
        fair_idx &= {i for i, c in enumerate(companies) if _looks_turkish(c)}
        log.info("Sadece Türk firma modu: %d firma kaldı", len(fair_idx))

    # --only-empty: zaten zenginleştirilmiş olanları atla
    if args.only_empty:
        targets = []
        skipped = []
        for i in sorted(fair_idx):
            c = companies[i]
            already_enriched = bool(
                c.get("website") or c.get("phone") or c.get("email")
            )
            if already_enriched:
                skipped.append(i)
            else:
                targets.append(i)
        log.info(
            "Sadece eksik mod: %d firma işlenecek, %d firma atlandı (zaten dolu)",
            len(targets), len(skipped),
        )
        if not targets:
            log.info("İşlenecek yeni firma yok. Bitti.")
            save_excel(companies)
            return
    else:
        targets = sorted(fair_idx)
        log.info(
            "Toplam %d firma, %d paralel thread ile zenginleştiriliyor",
            len(targets), MAX_WORKERS,
        )

    results: dict[int, dict] = {i: companies[i] for i in range(total)}
    failed: list[str] = []
    done_count = 0
    target_count = len(targets)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_idx = {
            executor.submit(enrich_one, dict(companies[i])): i
            for i in targets
        }
        for future in as_completed(future_to_idx):
            idx = future_to_idx[future]
            done_count += 1
            try:
                results[idx] = future.result()
                c = results[idx]
                has_data = any(c.get(k) for k in ("phone", "email", "linkedin", "website"))
                status = "✓" if has_data else "–"
                log.info("[%d/%d] %s %s | web:%s | tel:%s | mail:%s",
                         done_count, target_count, status, c["name"],
                         "✓" if c.get("website") else "–",
                         "✓" if c.get("phone") else "–",
                         "✓" if c.get("email") else "–")
            except Exception as e:
                results[idx] = companies[idx]
                failed.append(companies[idx]["name"])
                log.warning("[%d/%d] HATA %s: %s",
                            done_count, target_count, companies[idx]["name"], e)

            if done_count % 20 == 0:
                # Merge-tabanlı ara kayıt: sadece bulunan iletişim alanları
                # diske işlenir; Excel her ara kayıtta YAZILMAZ (10k satırlık
                # dosyayı 20 firmada bir yeniden üretmek dakikalar israf ediyordu
                # — kullanıcı istediği an webapp'ten export alabilir).
                ordered = [results.get(i, companies[i]) for i in range(total)]
                save_progress(ordered)
                log.info("Ara kayıt yapıldı (%d/%d)", done_count, target_count)

    ordered = [results.get(i, companies[i]) for i in range(total)]
    disk = save_progress(ordered, backup=True)
    save_excel(disk)

    # ── Google Maps fallback: telefonu hâlâ eksik olanları Maps'e gönder ──
    # --fair verildiyse sadece o fuarın firmaları Maps'e gönderilir (dict'ler
    # ordered ile paylaşımlı olduğu için mutasyon otomatik yansır).
    if not args.no_maps:
        maps_targets = [ordered[i] for i in sorted(fair_idx)] if args.fair else ordered
        run_maps_fallback(maps_targets)
        disk = save_progress(ordered, backup=True)
        save_excel(disk)
    else:
        log.info("--no-maps: Google Maps fallback atlandı (maps_only.py ile ayrı çalıştırın)")

    success_count = sum(
        1 for c in ordered if any(c.get(k) for k in ("phone", "email", "linkedin", "website"))
    )
    log.info("=" * 55)
    log.info(
        "ÖZET: %d firmadan %d'sinin iletişim bilgisi var (bu çalıştırmada %d hedef)",
        total, success_count, target_count,
    )
    if failed:
        log.info("Atlayanlar (%d): %s", len(failed), ", ".join(failed[:20]))
    log.info("Çıktı: %s", EXCEL_FILE)


if __name__ == "__main__":
    main()
