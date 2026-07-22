"""
Fuar Takvimi Toplayıcı — Aşama 1 (Adım A)
Türkiye'de 2026'da düzenlenecek tüm fuarları:
  - TOBB resmi Excel dosyasından (289 fuar)
  - Tüyap fuar takvimi sayfasından (kart yapısı)
çeker, output/fairs.json olarak kaydeder.

Bu dosya, ileride yazılacak ui.py (Streamlit) için ana veri kaynağı olacak.
"""

import json
import re
import logging
import urllib3
from io import BytesIO
from pathlib import Path
from datetime import datetime

import requests
import openpyxl
from bs4 import BeautifulSoup

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)
FAIRS_FILE = OUTPUT_DIR / "fairs.json"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "tr-TR,tr;q=0.9,en-US;q=0.8,en;q=0.7",
}

TOBB_EXCEL_URLS = [
    "https://www.tobb.org.tr/FuarlarMudurlugu/Documents/FUARLAR/2026fuartakvimi.xlsx",
    "https://www.tobb.org.tr/FuarlarMudurlugu/Documents/FUARLAR/2027fuartakvimi.xlsx",
]

TUYAP_URL = "https://www.tuyap.com.tr/fuar-takvimi"

# Sektör tahmini için anahtar kelime tablosu
SECTOR_KEYWORDS = {
    "İş Makineleri & İnşaat": ["komatek", "iş makine", "inşaat", "yapı", "beton",
                                 "vinç", "yapım", "construction", "machinery"],
    "Endüstri & Otomasyon": ["win eurasia", "win", "endüstri", "otomasyon", "imalat",
                              "metal işleme", "robot", "industrial", "automation"],
    "Otomotiv": ["automechanika", "otomotiv", "automotive", "araç", "yedek parça"],
    "Gıda & Tarım": ["food", "gıda", "tarım", "fooditek", "ipack", "agriculture",
                     "hayvancılık", "süt"],
    "Mobilya & İç Mekan": ["mobilya", "furniture", "ev tekstil", "dekorasyon"],
    "Tekstil & Moda": ["tekstil", "textile", "fashion", "moda", "konfeksiyon",
                        "kumaş", "deri"],
    "Sağlık & Medikal": ["medikal", "medical", "sağlık", "pharma", "ilaç",
                          "diş", "expomed"],
    "Eğitim": ["eğitim", "education", "kitap"],
    "Turizm & Otelcilik": ["turizm", "tourism", "otel", "horeca", "gastronomi"],
    "Kozmetik & Güzellik": ["beauty", "kozmetik", "güzellik"],
    "Enerji & Çevre": ["enerji", "energy", "solar", "rüzgar", "çevre", "geri dönüşüm",
                        "ifat"],
    "Savunma & Havacılık": ["idef", "savunma", "defence", "havacılık", "aerospace",
                              "drone"],
    "Bilişim & Teknoloji": ["bilişim", "teknoloji", "tech", "yazılım", "ai"],
    "Tarım Makineleri": ["agroexpo", "tarım maki", "ziraat"],
    "Ambalaj": ["ambalaj", "packaging", "ipack"],
    "Lojistik & Taşımacılık": ["logistik", "lojistik", "logistics", "taşımacılık"],
    "Yapı Malzemeleri": ["yapı malz", "isk", "ish", "tesisat", "doğalgaz"],
}


def guess_sector(name: str) -> str:
    lower = name.lower()
    for sector, keywords in SECTOR_KEYWORDS.items():
        if any(k in lower for k in keywords):
            return sector
    return "Genel / Diğer"


TR_MONTHS = {
    "ocak": 1, "şubat": 2, "mart": 3, "nisan": 4, "mayıs": 5, "haziran": 6,
    "temmuz": 7, "ağustos": 8, "eylül": 9, "ekim": 10, "kasım": 11, "aralık": 12,
}


def parse_date(value) -> str:
    """Excel hücresini ISO tarihe çevir (YYYY-MM-DD)."""
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    # "06.01.2026" → 2026-01-06
    m = re.match(r"^(\d{1,2})[./\-](\d{1,2})[./\-](\d{4})$", s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    # "06-09 Mayıs 2026" → 2026-05-06
    m = re.search(r"(\d{1,2})\s*[-–]\s*(\d{1,2})\s+([a-zçğıöşü]+)\s+(\d{4})",
                  s.lower())
    if m:
        d1, _, mo_name, y = m.groups()
        mo = TR_MONTHS.get(mo_name, 0)
        if mo:
            return f"{y}-{mo:02d}-{int(d1):02d}"
    # "12 Mayıs 2026" → 2026-05-12
    m = re.search(r"(\d{1,2})\s+([a-zçğıöşü]+)\s+(\d{4})", s.lower())
    if m:
        d, mo_name, y = m.groups()
        mo = TR_MONTHS.get(mo_name, 0)
        if mo:
            return f"{y}-{mo:02d}-{int(d):02d}"
    return ""


def fetch_tobb() -> list[dict]:
    """
    TOBB resmi Excel yapısı (R3 başlık satırı):
      C1=NO  C2=BAŞLAMA  C3=BİTİŞ  C4=FUARIN ADI  C5=KONUSU
      C6=ÜRÜN HİZMET GRUPLARI  C7=TÜRÜ  C8=YER  C9=ŞEHİR
      C10=DÜZENLEYİCİ  C14=WEB  C15=E-MAIL
    """
    fairs = []
    for url in TOBB_EXCEL_URLS:
        log.info("TOBB Excel indiriliyor: %s", url)
        try:
            r = requests.get(url, headers=HEADERS, timeout=60, verify=False)
            if r.status_code != 200:
                log.warning("TOBB %s → status %d", url, r.status_code)
                continue
            wb = openpyxl.load_workbook(BytesIO(r.content), data_only=True)
        except Exception as e:
            log.error("TOBB indirilemedi: %s", e)
            continue

        ws = wb.active
        log.info("Sheet: %s (%d satır)", ws.title, ws.max_row)

        # R3 başlık satırı, R4'ten itibaren veri
        for row_idx in range(4, ws.max_row + 1):
            no    = ws.cell(row_idx, 1).value
            start = parse_date(ws.cell(row_idx, 2).value)
            end   = parse_date(ws.cell(row_idx, 3).value)
            name  = str(ws.cell(row_idx, 4).value or "").strip()
            topic = str(ws.cell(row_idx, 5).value or "").strip().replace("_x000D_", "").strip()
            groups = str(ws.cell(row_idx, 6).value or "").strip().replace("_x000D_", "").strip()
            kind  = str(ws.cell(row_idx, 7).value or "").strip()
            venue = str(ws.cell(row_idx, 8).value or "").strip()
            city  = str(ws.cell(row_idx, 9).value or "").strip()
            organizer = str(ws.cell(row_idx, 10).value or "").strip().replace("_x000D_", "").strip()
            web   = str(ws.cell(row_idx, 14).value or "").strip()
            email = str(ws.cell(row_idx, 15).value or "").strip()

            if not name or len(name) < 5:
                continue

            # Web URL normalize
            if web and not web.startswith("http"):
                web = "https://" + web

            fair = {
                "no": no,
                "name": name,
                "start_date": start,
                "end_date": end,
                "city": city,
                "venue": venue,
                "topic": topic,
                "groups": groups,
                "kind": kind,
                "organizer": organizer,
                "url": web,
                "email": email,
                "sector": guess_sector(name + " " + topic + " " + groups),
                "source": "TOBB",
                "year": start[:4] if start else "",
                "participants_url": "",
            }
            fairs.append(fair)

    log.info("TOBB: %d fuar bulundu", len(fairs))
    return fairs


def fetch_tuyap() -> list[dict]:
    log.info("Tüyap fuar takvimi çekiliyor (Playwright): %s", TUYAP_URL)
    fairs = []
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            ctx = browser.new_context(
                user_agent=HEADERS["User-Agent"],
                locale="tr-TR",
                ignore_https_errors=True,
            )
            page = ctx.new_page()
            page.goto(TUYAP_URL, timeout=60_000, wait_until="networkidle")
            # Lazy-load için scroll
            for _ in range(8):
                page.keyboard.press("End")
                page.wait_for_timeout(800)
            html = page.content()
            browser.close()

        soup = BeautifulSoup(html, "lxml")
        seen = set()

        # Kart yapısı genelde "fuar" içeren class'lara sahip
        card_selectors = [
            "[class*='fuar-kart']", "[class*='card']", "[class*='event']",
            "article", "[class*='item']",
        ]

        for sel in card_selectors:
            for card in soup.select(sel):
                # İsim
                name_el = card.find(["h2", "h3", "h4", "h5", "strong"])
                name = name_el.get_text(strip=True) if name_el else ""
                if not name or len(name) < 5 or len(name) > 200:
                    continue
                if name in seen:
                    continue

                # Tarih ve şehir → tüm kart metni içinden regex
                full = card.get_text(separator=" ", strip=True)
                date_match = re.search(
                    r"(\d{1,2})\s*[-–]?\s*(\d{1,2})?\s*([a-zçğıöşü]+)\s+(\d{4})",
                    full.lower(),
                )
                start = ""
                if date_match:
                    start = parse_date(date_match.group(0))
                end = start

                # Şehir — büyük harfle başlayan il adlarını ara
                cities = re.findall(
                    r"\b(İstanbul|Ankara|İzmir|Bursa|Antalya|Adana|Gaziantep|Konya|"
                    r"Mersin|Diyarbakır|Kayseri|Eskişehir|Trabzon|Samsun|Denizli|Hatay)\b",
                    full,
                )
                city = cities[0] if cities else ""

                # URL
                a = card.find("a", href=True)
                url = a["href"] if a else ""
                if url and not url.startswith("http"):
                    url = "https://www.tuyap.com.tr" + url

                seen.add(name)
                fair = {
                    "name": name,
                    "start_date": start,
                    "end_date": end,
                    "city": city,
                    "topic": "",
                    "organizer": "Tüyap",
                    "address": "",
                    "sector": guess_sector(name),
                    "source": "Tüyap",
                    "year": start[:4] if start else "",
                    "url": url,
                    "participants_url": "",
                }
                fairs.append(fair)

        log.info("Tüyap: %d fuar bulundu", len(fairs))
    except Exception as e:
        log.error("Tüyap hatası: %s", e)
    return fairs


def merge_and_dedupe(fairs: list[dict]) -> list[dict]:
    """Aynı fuar (isim benzerliği + yakın tarih) tek kayıt olsun."""
    def normalize(name: str) -> str:
        return re.sub(r"[^\w]", "", name.lower())

    by_key: dict[str, dict] = {}
    for f in fairs:
        key = normalize(f["name"])[:40] + "|" + f.get("start_date", "")[:7]
        if key in by_key:
            existing = by_key[key]
            # Daha zengin kayıt kazanır (TOBB > Tüyap için topic/address)
            for k in ("city", "topic", "organizer", "address", "url", "participants_url"):
                if not existing.get(k) and f.get(k):
                    existing[k] = f[k]
            if existing["source"] != f["source"]:
                existing["source"] = existing["source"] + " + " + f["source"]
        else:
            by_key[key] = dict(f)
    return list(by_key.values())


def link_known_participants(fairs: list[dict]) -> list[dict]:
    """Bildiğimiz fuarlar için katılımcı listesi URL'sini ekle."""
    KNOWN = {
        "komatek": "https://komatekfuar.com/en/list-of-komatek-2026-participants/",
        "win eurasia": "https://platform.win-eurasia.com/participants?new",
    }
    for f in fairs:
        lower = f["name"].lower()
        for key, url in KNOWN.items():
            if key in lower:
                f["participants_url"] = url
                break
    return fairs


def main():
    all_fairs = []
    all_fairs += fetch_tobb()
    all_fairs += fetch_tuyap()

    log.info("Toplam ham kayıt: %d", len(all_fairs))
    fairs = merge_and_dedupe(all_fairs)
    fairs = link_known_participants(fairs)

    # Tarihe göre sırala
    fairs.sort(key=lambda f: (f.get("start_date") or "9999-99-99", f["name"]))

    FAIRS_FILE.write_text(
        json.dumps(fairs, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    log.info("=" * 55)
    log.info("Toplam %d fuar kaydedildi → %s", len(fairs), FAIRS_FILE)

    # Sektör dağılımı özet
    sector_counts: dict[str, int] = {}
    for f in fairs:
        sector_counts[f["sector"]] = sector_counts.get(f["sector"], 0) + 1
    log.info("Sektör dağılımı:")
    for sec, n in sorted(sector_counts.items(), key=lambda x: -x[1]):
        log.info("  %3d  %s", n, sec)


if __name__ == "__main__":
    main()
