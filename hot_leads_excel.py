"""
Sıcak arama listesi Excel'i üretir → output/SICAK_ARAMA_LISTESI.xlsx

Satış mantığı: fuarı önümüzdeki 150 gün içinde olan firmalar stand kararını şu
an veriyor. Skor = sıcaklık + marka gücü + yabancılık + iletişim varlığı
(datastore.compute_satis_skoru). Telefonu OLAN firmalar önde — bugün aranabilir.

Sheet'ler:
  1) "🔥 Bugün Ara"   — telefonu olan sıcak leadler, skora göre (ilk 300)
  2) "🌍 Yabancı Sıcak" — yabancı + sıcak (mail ile ulaşılacaklar dahil)
  3) "Fuar Özeti"      — yaklaşan fuar başına sıcak lead sayısı

Çalıştırma: python hot_leads_excel.py
"""

import logging
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

import datastore as ds

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

OUT_FILE = Path(__file__).parent / "output" / "SICAK_ARAMA_LISTESI.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="2F4F8F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HOT_FILL = PatternFill("solid", fgColor="FDEAEA")

COLS = ["Skor", "Fuar", "Fuar Tarihi", "Firma", "Ülke", "Menşei",
        "Telefon", "E-posta", "Website", "Öncelik", "Durum", "Not"]


def _write_sheet(wb, title, rows):
    ws = wb.create_sheet(title)
    ws.append(COLS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for c in rows:
        ws.append([
            c.get("satis_skoru", 0), c.get("fair", ""), c.get("fuar_tarihi", ""),
            c.get("name", ""), c.get("country", ""), c.get("mensei", ""),
            c.get("phone", ""), c.get("email", ""), c.get("website", ""),
            c.get("priority", ""), c.get("status", ""), c.get("note", ""),
        ])
    widths = [6, 34, 11, 42, 14, 10, 16, 28, 26, 8, 13, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


def main():
    data = ds.load()
    hot = [c for c in data if "🔥" in c.get("sicaklik", "")]
    hot.sort(key=lambda c: (-c.get("satis_skoru", 0), c.get("fuar_tarihi") or "9999"))

    callable_now = [c for c in hot if c.get("phone")][:300]
    foreign_hot = [c for c in hot if "🌍" in c.get("mensei", "")][:300]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "🔥 Bugün Ara", callable_now)
    _write_sheet(wb, "🌍 Yabancı Sıcak", foreign_hot)

    # Fuar özeti
    ws = wb.create_sheet("Fuar Özeti")
    ws.append(["Fuar", "Tarih", "Kalan Gün", "Sıcak Lead", "Telefonu Olan", "Yabancı"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    per_fair = {}
    today = date.today()
    for c in hot:
        f = c.get("fair", "")
        s = per_fair.setdefault(f, {"date": c.get("fuar_tarihi", ""), "n": 0, "tel": 0, "yab": 0})
        s["n"] += 1
        if c.get("phone"):
            s["tel"] += 1
        if "🌍" in c.get("mensei", ""):
            s["yab"] += 1
    for f, s in sorted(per_fair.items(), key=lambda kv: kv[1]["date"] or "9999"):
        days = (date.fromisoformat(s["date"]) - today).days if s["date"] else ""
        ws.append([f, s["date"], days, s["n"], s["tel"], s["yab"]])
    for i, w in enumerate([44, 11, 10, 11, 14, 9], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.active = 0
    wb.save(OUT_FILE)
    log.info("Yazıldı: %s | bugun-ara=%d, yabanci-sicak=%d, fuar=%d",
             OUT_FILE, len(callable_now), len(foreign_hot), len(per_fair))


if __name__ == "__main__":
    main()
